"""
Eva 知識庫更新腳本 v2
----------------------
每次 JID/牌價 的牌價表更新後，在本地執行這個腳本，
重新產生 eva_knowledge.txt（含品牌型號 + 技術知識 + 選型指南）。

執行方式：
    cd /Users/user/Downloads/Peggy_agent/line_bot
    python3 build_eva_knowledge.py

執行完後 commit + push，再到 Render 手動 Deploy。
"""

import os, glob, re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl

try:
    import xlrd
except ImportError:
    os.system("pip install xlrd -q")
    import xlrd

JID_PRICE_DIR = "/Users/user/Library/CloudStorage/GoogleDrive-s9220320@gmail.com/我的雲端硬碟/JID/牌價"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "eva_knowledge.txt")

# ── 找最新牌價檔案 ────────────────────────────────────────────────

def find_latest(pattern: str) -> str | None:
    files = glob.glob(os.path.join(JID_PRICE_DIR, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def extract_date_from_filename(path: str) -> str:
    name = os.path.basename(path)
    m = re.search(r'(\d{4}-\d{1,2}-\d{1,2}|\d{8})', name)
    if m:
        return m.group(1)
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d")

# ── 讀 .xls — 智慧偵測型號欄位 ────────────────────────────────────

def _find_model_col(ws, nrows: int) -> int:
    """偵測標題列，找「型號」或「品名」所在欄位（0-indexed）。預設 col 0。"""
    for ri in range(min(4, nrows)):
        row = [str(v).strip() for v in ws.row_values(ri)]
        for ci, cell in enumerate(row):
            if cell in ("型號",):
                return ci
        # 「品   名」或「品名」— 第一欄若是空的才用它
        for ci, cell in enumerate(row):
            if cell in ("品   名", "品名"):
                # 確認 col 0 是否全為空（代表型號確實在 ci）
                sample = [str(ws.cell_value(r, 0)).strip() for r in range(ri + 1, min(ri + 6, nrows))]
                if all(s == "" for s in sample):
                    return ci
                return 0
    return 0


def read_xls_brands_models(path: str) -> dict[str, list[str]]:
    wb = xlrd.open_workbook(path)
    result: dict[str, list[str]] = {}

    skip_sheets = {
        "Sheet1", "Power supply #1", "VS日幣", "mViz software",
        "Basler(封存)", "Basler凌華價格", "非主流產品", "PV500&PV200",
        "Saber1(IMPERX)", "CCS BTC",
    }
    # 工作表名稱 → 顯示名稱（方便後續輸出）
    display_name_map = {
        "Basler (ACE)": "Basler ACE",
        "Basler nonACE": "Basler (非ACE)",
        "蘇映視(new)": "蘇映視 INSNEX",
    }

    for sheet_name in wb.sheet_names():
        if sheet_name in skip_sheets:
            continue
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows < 3:
            continue

        model_col = _find_model_col(ws, ws.nrows)
        models = []

        for i in range(ws.nrows):
            raw = ws.row_values(i)
            if model_col >= len(raw):
                continue
            val = str(raw[model_col]).strip()
            if not val or len(val) < 2:
                continue
            # 跳過標題、空值、純數字
            if val in ("型號", "品   名", "品名", "ACE SERIES", "GIGE", "USB3", "CXP",
                       "Sony Camera", "Part Number", "ProNo", "Neurocle", "mViz 2.x List Price",
                       "LINE SCAN SERIES", "LINE SCAN", "LINESCANE"):
                continue
            if val.startswith(("●", "※", "#", "(", '"', "DIGITAL", "ANALOG", "FRAME",
                                "Camera Line", "Sprint", "Basler", "GimaGO", "P. O.")):
                continue
            if re.match(r'^[\d,\.\s]+$', val):
                continue
            models.append(val)

        if models:
            display = display_name_map.get(sheet_name, sheet_name)
            result[display] = models

    return result

# ── 讀 .xlsx — 智慧偵測型號欄位 ─────────────────────────────────────

def _find_model_col_xlsx(ws) -> int:
    for ri in range(min(4, ws.max_row or 1)):
        row = [str(c.value).strip() if c.value is not None else "" for c in list(ws.rows)[ri]]
        for ci, cell in enumerate(row):
            if cell == "型號":
                return ci
        for ci, cell in enumerate(row):
            if cell in ("品   名", "品名"):
                sample_col0 = []
                for r in list(ws.rows)[ri + 1: ri + 6]:
                    v = r[0].value
                    sample_col0.append(str(v).strip() if v is not None else "")
                if all(s == "" for s in sample_col0):
                    return ci
                return 0
    return 0


def read_xlsx_brands_models(path: str) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result: dict[str, list[str]] = {}
    skip_sheets = {"Sheet1", "schneider LENS Selection"}

    for name in wb.sheetnames:
        if name in skip_sheets:
            continue
        ws = wb[name]
        if ws.max_row is None or ws.max_row < 3:
            continue

        rows = list(ws.iter_rows(values_only=True))
        model_col = 0
        for ri in range(min(4, len(rows))):
            row = [str(v).strip() if v is not None else "" for v in rows[ri]]
            found = False
            for ci, cell in enumerate(row):
                if cell == "型號":
                    model_col = ci
                    found = True
                    break
            if found:
                break
            for ci, cell in enumerate(row):
                if cell in ("品   名", "品名"):
                    sample = [str(rows[r][0]).strip() if rows[r][0] is not None else ""
                              for r in range(ri + 1, min(ri + 6, len(rows)))]
                    model_col = ci if all(s in ("", "None") for s in sample) else 0
                    found = True
                    break
            if found:
                break

        models = []
        for row in rows:
            if model_col >= len(row):
                continue
            val = str(row[model_col]).strip() if row[model_col] is not None else ""
            if not val or len(val) < 2 or val == "None":
                continue
            if val in ("型號", "品   名", "品名", "Telecentric Lens", "品名/規格",
                       "CCS Lens+LED", "Detail", "原型號"):
                continue
            if val.startswith(("●", "※", "#", "(", '"', "Mega", "CCTV", "VMV")):
                continue
            if re.match(r'^[\d,\.\s]+$', val):
                continue
            models.append(val)

        if models:
            result[name] = models

    wb.close()
    return result

# ── 技術知識庫（直接寫入，來源：Staff Training + 內部教材）──────────

TECH_KNOWLEDGE = """
## 二、機器視覺技術知識

### 2-1 工業相機基礎

**感測器類型**
- CMOS：現代主流，低成本、低耗電、高幀率，適合大多數應用
- CCD：影像雜訊低、成像品質穩定，適合高精度量測（現已逐漸被高階 CMOS 取代）

**傳輸介面比較**
| 介面 | 頻寬 | 最大線長 | 特色 |
|------|------|---------|------|
| GigE | 100 MB/s | 100m | 長距離首選，支援 PoE 單線供電，多機布建成本低 |
| USB 3.0 | 350 MB/s | 8m | 隨插即用，低 CPU 負載，短距離高頻寬首選 |
| Camera Link | 850 MB/s | 10m | 需擷取卡，高速高精度場合 |
| CoaXPress (CXP-12) | 12.5 Gbps | 40m | 超高速，一線整合影像+電力+控制 |

**常見相機類型**
- Area Scan（面掃）：最常見，拍一張完整影像，適合靜態或慢速物件
- Line Scan（線掃）：一次拍一條線，適合捲狀物（薄膜、布料、PCB）連續檢測
- 3D 相機：輸出點雲或高度圖，適合量測體積、高度差（如 深視 SR 系列）

### 2-2 鏡頭基礎

**核心參數**
- **焦距（mm）**：數字越小視角越廣，數字越大視野越窄（但能看更遠）
- **F值（光圈）**：F值越小進光量越多、景深越淺；F值越大景深越深
- **工作距離（WD）**：鏡頭到待測物的距離
- **FOV（視野範圍）**：相機能看到的實際範圍（mm）

**FOV 計算公式**
FOV = 感測器尺寸 ÷ 鏡頭倍率
（例：感測器 6.4mm，倍率 0.1x → FOV = 64mm）

**Mount 類型**
- C-mount：最常見，工業相機標準
- CS-mount：需加 5mm 延伸環才能對焦
- S-mount（M12）：最小，用於嵌入式場合

**鏡頭種類**
- CCTV Lens：一般鏡頭，最泛用
- Telecentric Lens（遠心鏡頭）：畸變極小，適合精密尺寸量測
- Macro Lens（微距鏡頭）：近距離高放大倍率
- 電動變焦鏡頭：可遠端調整焦距

### 2-3 光源基礎

**LED 顏色與波長對應**
| 顏色 | 波長 | 用途 |
|------|------|------|
| 紅光 | 660nm | 工業最常用，成像穩定 |
| 藍光 | 470nm | 凸顯金屬刮痕、細微特徵 |
| 綠光 | 525nm | 特定材質顯色 |
| 白光 | 複合 | 一般通用，色溫建議 5500K |
| 近紅外 | 900nm+ | 穿透特定材質、排除環境光干擾 |

**打光原理：同色反射、異色吸收**
- 欲讓特徵變亮（白）→ 用相同顏色光源
- 欲讓特徵變暗（黑）→ 用互補色光源
- 例：紅色文字在橘色背景上 → 用藍光讓紅字變黑突顯

**光源類型與適用場景**
| 光源類型 | 適用場景 |
|---------|---------|
| 環形光（Ring Light）| 強化輪廓、邊緣量測、一般外觀檢測 |
| 低角度環形光 | 突顯表面刮痕、浮雕文字、電池/藥錠邊緣 |
| 同軸光（Coaxial）| 鏡面材質（CPU針腳、金屬銘板、玻璃），消除反光 |
| 穹頂光（Dome）| 曲面零件（螺帽、圓形件），消除陰影，均勻照明 |
| 背光（Backlight）| 輪廓尺寸量測（Silhouette）、透明瓶液位檢測 |
| 條型光（Bar/Line）| 大面積平面、Line scan 相機配套光源 |

**明視野 vs 暗視野**
- 明視野（Bright Field）：直接光反射，背景亮、文字暗
- 暗視野（Dark Field）：低角度照射，凸起特徵發亮、背景暗

## 三、選型計算公式

### 3-1 精度計算（3 倍像素規則）
```
Pixel size = FOV ÷ 解析度（像素數）
實務精度 = Pixel size × 3
```
**範例**：FOV 40mm，相機 2592 px
→ Pixel size = 40 ÷ 2592 = 0.0154 mm/px
→ 實務精度 ≈ 0.046 mm

### 3-2 頻寬計算
```
頻寬 (bps) = 水平解析度 × 垂直解析度 × 位元深度 × 幀率
```
**範例**：1280×1024，8bit，60fps
→ 1280 × 1024 × 8 × 60 = 629,145,600 bps ≈ 75 MB/s
→ USB 3.0（350 MB/s）可支援，GigE（100 MB/s）不夠

### 3-3 頻寬安全原則
- 單一匯流排總頻寬不超過介面上限的 80%
- GigE 每台相機建議 ≤ 80 MB/s（留 20% 餘量）

## 四、應用場景推薦

### 4-1 AOI 外觀檢測（瑕疵、刮痕、異物）
- 相機：Area scan，根據速度選 GigE 或 USB3
- 光源：同軸光（鏡面）或低角度環形光（霧面/粗糙面）
- 鏡頭：Telecentric（需精密量測）或 CCTV Lens（一般檢測）

### 4-2 尺寸量測（精密）
- 相機：高解析度（5MP 以上），全域快門
- 鏡頭：Telecentric Lens（無畸變），正確 WD
- 光源：背光（輪廓量測）或同軸光（平面量測）
- 注意：先用公式確認 pixel size × 3 ≤ 公差需求

### 4-3 PCB / 電子元件檢測
- 相機：高解析，Color 或 Mono 視需求
- 光源：同軸光（PIN 腳反光面）＋ 環形光補充
- 鏡頭：配合 WD 選擇適當倍率

### 4-4 捲狀物 / 大面積連續檢測
- 相機：Line scan（racer 系列）
- 光源：條型光（Line 光源），與掃描方向垂直
- 鏡頭：Line scan 專用鏡頭

### 4-5 3D 量測（高度差、體積）
- 相機：深視 SR 系列（3D 輪廓掃描儀）或 Basler blaze（ToF）
- SR 系列特色：整合藍色雷射、超高速（8000 輪廓/秒）、IP67 防護

### 4-6 透明物件 / 液位檢測
- 光源：背光（Backlight）
- 相機：Area scan，依速度選介面

## 五、我們代理品牌的特色

**Basler（德國，相機市占第一）**
- 三年保固、德國原廠技術支援
- 產品線：ace（通用）、boost（CXP高速）、dart（嵌入式低成本）、racer（Line scan）、blaze（3D ToF）
- ace 2 Pro 功能：Compression Beyond（GigE 突破頻寬限制）、Pixel Beyond（調整感測器特性）

**CCS（日本，光源高端品牌）**
- 兩年光源亮度保固
- 日本機器視覺光源研發先驅，種類最多
- 含 LDR（環形）、LDL（條型）、LFV（同軸）、LAI（背光）等系列

**深視 SSZN（SR 系列 3D 感測器）**
- 整合藍色雷射，無需外部光源
- 超高動態範圍（一般相機的 100 倍），黑色到金屬鏡面都能量測
- IP67 防護，工廠環境適用

**MOZI 軟體**
- JIDIEN 自有視覺分析軟體
- 支援 2D 瑕疵檢測、手臂導引（VGR）、3D 量測
- 深度學習模組（AI 瑕疵分類）

**Bitflow（擷取卡）**
- 美國品牌，高速影像擷取卡
- 支援 Camera Link 和 CoaXPress 兩種介面
"""

# ── 產生知識庫 ────────────────────────────────────────────────────

def build_knowledge(cam_path: str | None, lens_path: str | None) -> str:
    lines = []
    today = datetime.now().strftime("%Y-%m-%d")
    lines.append("# JIDIN_Peggy 產品知識庫（Eva 專用）")
    lines.append(f"# 更新日期：{today}")
    lines.append("# 用途：Eva 回答客戶詢問，含品牌代理查詢、技術知識、選型建議")
    lines.append("")
    lines.append("## 一、代理品牌與型號")
    lines.append("")

    if cam_path:
        cam_date = extract_date_from_filename(cam_path)
        lines.append(f"### 工業相機 & 光源 & 周邊（牌價表：{cam_date}）")
        lines.append("")
        brands_models = read_xls_brands_models(cam_path)
        for brand, models in brands_models.items():
            lines.append(f"**{brand}**")
            for m in models[:40]:
                lines.append(f"  - {m}")
            if len(models) > 40:
                lines.append(f"  （共 {len(models)} 筆型號，以上為部分列舉）")
            lines.append("")
    else:
        lines.append("### 工業相機（牌價表未找到）")
        lines.append("")

    if lens_path:
        lens_date = extract_date_from_filename(lens_path)
        lines.append(f"### 鏡頭（牌價表：{lens_date}）")
        lines.append("")
        brands_models = read_xlsx_brands_models(lens_path)
        for brand, models in brands_models.items():
            lines.append(f"**{brand}**")
            for m in models[:40]:
                lines.append(f"  - {m}")
            if len(models) > 40:
                lines.append(f"  （共 {len(models)} 筆型號，以上為部分列舉）")
            lines.append("")
    else:
        lines.append("### 鏡頭（牌價表未找到）")
        lines.append("")

    # 加入技術知識
    lines.append(TECH_KNOWLEDGE)

    return "\n".join(lines)

# ── 主程式 ────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("正在尋找最新牌價表...")
    cam_path = find_latest("(業務)Vision牌價表[0-9]*.xls")
    if not cam_path:
        cam_path = find_latest("*牌價表[0-9]*.xls")
    print(f"  相機牌價表：{os.path.basename(cam_path) if cam_path else '未找到'}")

    lens_path = find_latest("*鏡頭*.xlsx")
    print(f"  鏡頭牌價表：{os.path.basename(lens_path) if lens_path else '未找到'}")

    print("正在抽取品牌/型號與技術知識...")
    content = build_knowledge(cam_path, lens_path)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(content)

    lines_count = content.count("\n")
    size = len(content.encode("utf-8"))
    print(f"完成！eva_knowledge.txt 已更新（{lines_count} 行，{size/1024:.1f} KB）")
    print(f"路徑：{OUTPUT_PATH}")
    print()
    print("下一步：")
    print("  git add eva_knowledge.txt")
    print("  git commit -m 'update: Eva knowledge base v2'")
    print("  git push origin main")
    print("  → Render 手動 Deploy latest commit")
